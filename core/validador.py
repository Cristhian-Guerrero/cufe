"""
═══════════════════════════════════════════════════════════════════════════
VALIDADOR DE CUFES - CUFE DIAN AUTOMATION v3.2.0
Valida formato, elimina duplicados y detecta CUFEs inválidos
Soporta NIT asociado: columna B en Excel, o "CUFE,NIT" / "CUFE\tNIT" en .txt
Retorna List[Tuple[str, str]] = [(cufe, nit), ...]
CUFEs sin NIT se reportan como inválidos con razón 'sin_nit'
═══════════════════════════════════════════════════════════════════════════
"""

import os
import re
from typing import List, Tuple, Dict
from utils import log

# Importación condicional de openpyxl (ya está en requirements.txt)
try:
    from openpyxl import load_workbook
    EXCEL_DISPONIBLE = True
except ImportError:
    EXCEL_DISPONIBLE = False
    log(0, "⚠️  openpyxl no disponible - archivos Excel deshabilitados", "WARN")


class ValidadorCUFE:
    """
    Validador de CUFEs con detección de errores
    
    Validaciones:
    - Longitud correcta (96 caracteres)
    - Formato hexadecimal válido
    - Eliminación de duplicados
    - Archivo de entrada existe
    
    Formatos soportados:
    - .txt (un CUFE por línea)
    - .xlsx (CUFEs en columna A)
    """
    
    LONGITUD_CUFE = 96
    
    def __init__(self):
        """Inicializa el validador"""
        self.cufes_validos = []
        self.cufes_invalidos = []
        self.duplicados_eliminados = 0
    
    def validar_archivo(self, ruta_archivo: str) -> Tuple[bool, str]:
        """
        Valida que el archivo existe y es legible
        
        Args:
            ruta_archivo: Ruta del archivo de CUFEs
            
        Returns:
            (existe, mensaje_error)
        """
        if not os.path.exists(ruta_archivo):
            return False, f"❌ Archivo no encontrado: {ruta_archivo}"
        
        if not os.path.isfile(ruta_archivo):
            return False, f"❌ No es un archivo: {ruta_archivo}"
        
        if not os.access(ruta_archivo, os.R_OK):
            return False, f"❌ Sin permisos de lectura: {ruta_archivo}"
        
        # Verificar extensión soportada
        extension = os.path.splitext(ruta_archivo)[1].lower()
        if extension not in ['.txt', '.xlsx']:
            return False, f"❌ Formato no soportado: {extension} (use .txt o .xlsx)"
        
        # Si es Excel, verificar que openpyxl esté disponible
        if extension == '.xlsx' and not EXCEL_DISPONIBLE:
            return False, "❌ Para archivos Excel instale: pip install openpyxl"
        
        return True, ""
    
    def es_cufe_valido(self, cufe: str) -> Tuple[bool, str]:
        """
        Valida que un CUFE tenga formato correcto
        
        Args:
            cufe: CUFE a validar
            
        Returns:
            (es_valido, razon_si_invalido)
        """
        if not cufe:
            return False, "vacío"
        
        # Verificar longitud
        if len(cufe) != self.LONGITUD_CUFE:
            return False, f"longitud incorrecta ({len(cufe)} chars, esperados {self.LONGITUD_CUFE})"
        
        # Verificar que sea hexadecimal (solo letras a-f y números 0-9)
        if not re.match(r'^[a-fA-F0-9]+$', cufe):
            return False, "formato no hexadecimal"
        
        return True, ""
    
    def _leer_archivo_txt(self, ruta_archivo: str) -> Tuple[List[Tuple[str, str]], str]:
        """
        Lee pares (CUFE, NIT) desde archivo .txt.

        Formatos soportados (un par por línea):
          - "CUFE,NIT"    separado por coma
          - "CUFE\tNIT"   separado por tabulador
          - "CUFE"        sin NIT (se registra como sin_nit)
        """
        try:
            with open(ruta_archivo, 'r', encoding='utf-8') as f:
                lineas = [l.strip() for l in f if l.strip()]

            if not lineas:
                return [], "❌ Archivo vacío"

            pares = []
            for linea in lineas:
                if '\t' in linea:
                    partes = linea.split('\t', 1)
                elif ',' in linea:
                    partes = linea.split(',', 1)
                else:
                    partes = [linea]

                cufe = partes[0].strip()
                nit  = partes[1].strip() if len(partes) > 1 else ''
                pares.append((cufe, nit))

            log(0, f"📄 Leídas {len(pares)} líneas desde archivo .txt", "INFO")
            return pares, ""

        except Exception as e:
            return [], f"❌ Error leyendo archivo .txt: {str(e)}"
    
    def _leer_archivo_excel(self, ruta_archivo: str) -> Tuple[List[Tuple[str, str]], str]:
        """
        Lee pares (CUFE, NIT) desde archivo .xlsx.

        Formato esperado:
        ┌─────────────────────────────┬───────────┐
        │ A (CUFE)                    │ B (NIT)   │
        ├─────────────────────────────┼───────────┤
        │ CUFE                        │ NIT       │  <- Encabezado (opcional)
        │ 8d528789cb2547d0...         │ 900123456 │
        └─────────────────────────────┴───────────┘
        NIT en col B es opcional; si falta se reporta como sin_nit.
        """
        if not EXCEL_DISPONIBLE:
            return [], "❌ openpyxl no disponible"

        try:
            wb = load_workbook(ruta_archivo, read_only=True, data_only=True)
            ws = wb.active
            log(0, f"📊 Leyendo Excel: hoja '{ws.title}'", "INFO")

            pares = []
            primera_fila = True

            for row in ws.iter_rows(min_col=1, max_col=2, values_only=True):
                val_cufe = row[0]
                val_nit  = row[1] if len(row) > 1 else None

                if val_cufe is None:
                    continue

                cufe_str = str(val_cufe).strip()
                if not cufe_str:
                    continue

                # Detectar fila de encabezado
                if primera_fila:
                    primera_fila = False
                    if cufe_str.upper() in ['CUFE', 'CUFES', 'CÓDIGO', 'CODIGO']:
                        log(0, f"  ✓ Encabezado detectado en fila 1", "INFO")
                        continue

                nit_str = str(val_nit).strip() if val_nit is not None else ''
                pares.append((cufe_str, nit_str))

            wb.close()

            if not pares:
                return [], "❌ No se encontraron CUFEs en columna A"

            log(0, f"📄 Leídos {len(pares)} pares CUFE/NIT desde Excel", "INFO")
            return pares, ""

        except Exception as e:
            return [], f"❌ Error leyendo archivo Excel: {str(e)}"
    
    def cargar_y_validar(self, ruta_archivo: str, eliminar_duplicados: bool = True) -> Tuple[List[Tuple[str, str]], Dict[str, any]]:
        """
        Carga y valida pares (CUFE, NIT) desde archivo (.txt o .xlsx).

        Returns:
            (lista_pares_validos, estadisticas)
            donde cada par es (cufe: str, nit: str)
        """
        existe, error = self.validar_archivo(ruta_archivo)
        if not existe:
            log(0, error, "ERROR")
            return [], {'error': error}

        extension = os.path.splitext(ruta_archivo)[1].lower()

        if extension == '.txt':
            pares, error = self._leer_archivo_txt(ruta_archivo)
        elif extension == '.xlsx':
            pares, error = self._leer_archivo_excel(ruta_archivo)
        else:
            error = f"❌ Extensión no soportada: {extension}"
            log(0, error, "ERROR")
            return [], {'error': error}

        if error:
            log(0, error, "ERROR")
            return [], {'error': error}

        if not pares:
            error = "❌ No se encontraron datos"
            log(0, error, "ERROR")
            return [], {'error': error}

        cufes_vistos = set()
        self.cufes_validos = []
        self.cufes_invalidos = []
        self.duplicados_eliminados = 0

        for idx, (cufe, nit) in enumerate(pares, 1):
            if cufe in cufes_vistos:
                self.duplicados_eliminados += 1
                if eliminar_duplicados:
                    log(0, f"⚠️  CUFE #{idx} duplicado (ignorado)", "WARN")
                    continue
                else:
                    log(0, f"ℹ️  CUFE #{idx} duplicado (conservado)", "INFO")

            es_valido, razon = self.es_cufe_valido(cufe)

            if not es_valido:
                self.cufes_invalidos.append({'linea': idx, 'cufe': cufe, 'razon': razon})
                log(0, f"⚠️  CUFE #{idx} inválido: {razon}", "WARN")
                continue

            if not nit:
                self.cufes_invalidos.append({'linea': idx, 'cufe': cufe, 'razon': 'sin_nit'})
                log(0, f"⚠️  CUFE #{idx} sin NIT — no se puede procesar", "WARN")
                continue

            self.cufes_validos.append((cufe, nit))
            cufes_vistos.add(cufe)

        stats = {
            'total_lineas': len(pares),
            'validos': len(self.cufes_validos),
            'invalidos': len(self.cufes_invalidos),
            'duplicados': self.duplicados_eliminados,
            'formato': extension
        }

        self._mostrar_resumen(stats)
        return self.cufes_validos, stats
    
    def _mostrar_resumen(self, stats: Dict):
        """Muestra resumen de validación"""
        print("\n" + "="*70)
        print("📋 VALIDACIÓN DE CUFEs")
        print("="*70)
        
        log(0, f"📄 Formato: {stats.get('formato', 'desconocido').upper()}", "INFO")
        log(0, f"📄 Total líneas: {stats['total_lineas']}", "INFO")
        log(0, f"✅ CUFEs válidos: {stats['validos']}", "OK")
        
        if stats['duplicados'] > 0:
            log(0, f"🔄 Duplicados eliminados: {stats['duplicados']}", "WARN")
        
        if stats['invalidos'] > 0:
            log(0, f"❌ CUFEs inválidos: {stats['invalidos']}", "ERROR")
            
            # Mostrar primeros 5 inválidos
            for item in self.cufes_invalidos[:5]:
                log(0, f"   Línea {item['linea']}: {item['razon']}", "ERROR")
            
            if len(self.cufes_invalidos) > 5:
                log(0, f"   ... y {len(self.cufes_invalidos) - 5} más", "ERROR")
        
        print("="*70 + "\n")
        
        # Si no hay CUFEs válidos, es un error crítico
        if stats['validos'] == 0:
            log(0, "❌ No hay CUFEs válidos para procesar", "CRIT")
            return False
        
        return True


# ═══════════════════════════════════════════════════════════════════════════
# FUNCIÓN DE CONVENIENCIA (compatibilidad con código original)
# ═══════════════════════════════════════════════════════════════════════════

def cargar_cufes(ruta_archivo: str, eliminar_duplicados: bool = True, invalidos_out: list = None) -> List[Tuple[str, str]]:
    """
    Carga y valida pares (CUFE, NIT) desde archivo .txt o .xlsx.

    Args:
        ruta_archivo: Ruta del archivo (.txt o .xlsx)
        eliminar_duplicados: Si False, conserva duplicados
        invalidos_out: Lista opcional; se rellena con dicts {'cufe', 'razon', 'linea'}

    Returns:
        Lista de tuplas [(cufe, nit), ...] listas para procesar
    """
    validador = ValidadorCUFE()
    pares, stats = validador.cargar_y_validar(ruta_archivo, eliminar_duplicados=eliminar_duplicados)

    if invalidos_out is not None:
        invalidos_out.extend(validador.cufes_invalidos)

    if stats.get('error') or stats.get('validos', 0) == 0:
        return []

    return pares


# ═══════════════════════════════════════════════════════════════════════════
# TESTING
# ═══════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    print("\n" + "="*70)
    print("🧪 PROBANDO VALIDADOR DE CUFEs v3.1.0")
    print("="*70 + "\n")
    
    import sys
    
    if len(sys.argv) > 1:
        archivo = sys.argv[1]
        print(f"📄 Validando: {archivo}\n")
        
        cufes = cargar_cufes(archivo)
        
        print(f"\n✅ Resultado: {len(cufes)} CUFEs válidos listos para procesar")
    else:
        print("ℹ️  Uso: python3 -m core.validador <archivo_cufes.txt|.xlsx>")
        
        # Test con datos de ejemplo
        print("\n🧪 Test con CUFEs de ejemplo (.txt):\n")
        
        # Crear archivo temporal de prueba
        with open('/tmp/test_cufes.txt', 'w') as f:
            f.write("8d528789cb2547d090fcae2acccfc04e8588a8b80ea7de7bda94ff33c11a0ccf564fa6c73c49ca8ca43809557eaa3a3a\n")  # Válido
            f.write("8d528789cb2547d090fcae2acccfc04e8588a8b80ea7de7bda94ff33c11a0ccf564fa6c73c49ca8ca43809557eaa3a3a\n")  # Duplicado
            f.write("INVALIDO123\n")  # Inválido (muy corto)
            f.write("gggggggggggggggggggggggggggggggggggggggggggggggggggggggggggggggggggggggggggggggggggggggggggggg\n")  # Inválido (g no es hex)
            f.write("c64601b7a5b65b68d2ff2a7e7c31d48beb674bc0974172368193825144d5aa5c1d7d444412ae5fbc7819e3e70341bc71\n")  # Válido
        
        cufes = cargar_cufes('/tmp/test_cufes.txt')
        print(f"\n✅ {len(cufes)} CUFEs válidos encontrados en .txt")
    
    print("\n" + "="*70 + "\n")
