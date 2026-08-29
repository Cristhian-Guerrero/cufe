"""
═══════════════════════════════════════════════════════════════════════════
GENERADOR DE EXCEL - CUFE DIAN AUTOMATION (FULL DATA)
v6.4.1 - Base gravable por tarifa (0%/5%/19%) + Total Base
         (Cuadre IVA se calcula y se loguea pero no se muestra en columna)
═══════════════════════════════════════════════════════════════════════════
"""

import os
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from utils import log

class GeneradorExcel:
    
    # ESTRUCTURA DEFINITIVA Y COMPLETA
    # Tupla: (Key_Diccionario, Titulo_Excel, Ancho_Columna, ID_Grupo)
    COLUMNAS_DEF = [
        # GRUPO 0: CONTROL
        ('Numero', 'N°', 6, 0), 
        ('Estado', 'Estado', 12, 0),
        ('Eventos', 'Eventos RADIAN', 20, 0),  # <--- NUEVO
        ('Numero_Factura', 'Factura N°', 15, 0),
        
        # GRUPO 1: DATOS DEL DOCUMENTO
        ('CUFE', 'CUFE', 25, 1), 
        ('Fecha_Emision', 'Emisión', 12, 1), 
        ('Fecha_Vencimiento', 'Vencimiento', 12, 1), 
        ('Tipo_Operacion', 'Tipo Operación', 18, 1), 
        ('Forma_Pago', 'Forma Pago', 15, 1), 
        ('Medio_Pago', 'Medio Pago', 20, 1), 
        
        # GRUPO 2: EMISOR (VENDEDOR)
        ('Emisor_RazonSocial', 'Razón Social Vendedor', 35, 2), 
        ('Emisor_NombreComercial', 'Nombre Comercial', 25, 2),
        ('Emisor_NIT', 'NIT Vendedor', 15, 2),
        ('Emisor_Pais', 'País', 12, 2),                  # <--- NUEVO
        ('Emisor_Departamento', 'Depto.', 15, 2), 
        ('Emisor_Municipio', 'Ciudad', 15, 2), 
        ('Emisor_Direccion', 'Dirección', 30, 2),
        ('Emisor_Telefono', 'Teléfono', 15, 2),          # <--- NUEVO
        ('Emisor_Correo', 'Email', 25, 2),               # <--- NUEVO
        ('Emisor_ActividadEconomica', 'Actividad', 20, 2), # <--- NUEVO
        ('Emisor_RegimenFiscal', 'Régimen', 15, 2),
        
        # GRUPO 3: ADQUIRIENTE (CLIENTE)
        ('Adq_Tipo', 'Tipo Pers.', 10, 3), 
        ('Adq_NumeroDocumento', 'NIT / CC Cliente', 15, 3), 
        ('Adq_RazonSocial', 'RAZÓN SOCIAL (LEGAL)', 35, 3), 
        ('Adq_NombreComercial', 'Establecimiento', 25, 3),
        ('Adq_Pais', 'País', 12, 3),                     # <--- NUEVO
        ('Adq_Departamento', 'Depto.', 15, 3), 
        ('Adq_Municipio', 'Ciudad', 15, 3), 
        ('Adq_Direccion', 'Dirección', 30, 3), 
        ('Adq_Telefono', 'Teléfono', 15, 3),             # <--- NUEVO
        ('Adq_Correo', 'Email', 25, 3),
        ('Adq_Responsabilidad', 'Resp. Tributaria', 20, 3), # <--- NUEVO
        ('Adq_RegimenFiscal', 'Régimen', 15, 3),         # <--- NUEVO
        
        # GRUPO 4: FINANCIERO
        ('Subtotal', 'Subtotal', 16, 4),
        ('Total_Bruto', 'Total Bruto', 16, 4),
        ('Base_0', 'Base IVA 0%', 14, 4),
        ('Base_5', 'Base IVA 5%', 14, 4),
        ('IVA_5', 'IVA 5%', 14, 4),
        ('Base_19', 'Base IVA 19%', 14, 4),
        ('IVA_19', 'IVA 19%', 14, 4),
        ('INC', 'INC', 14, 4),
        ('Bolsas', 'Bolsas', 12, 4),
        ('Total_Base', 'Total Base', 16, 4),
        ('Total_Factura', 'TOTAL A PAGAR', 18, 4),
        ('Anticipos', 'Anticipos', 14, 4),
        ('Rete_Fuente', 'ReteFuente', 14, 4),
        ('Rete_IVA', 'ReteIVA', 14, 4),
        ('Rete_ICA', 'ReteICA', 14, 4),
        
        # GRUPO 5: GESTIÓN
        ('Ruta_PDF', 'Soporte', 12, 5), 
        ('Notas', 'Observaciones', 30, 5)
    ]

    GRUPOS_INFO = {
        0: {'titulo': 'CONTROL', 'color': '404040', 'texto': 'FFFFFF'},
        1: {'titulo': 'DATOS DEL DOCUMENTO', 'color': '1F4E78', 'texto': 'FFFFFF'},
        2: {'titulo': 'DATOS DEL EMISOR', 'color': '375623', 'texto': 'FFFFFF'},
        3: {'titulo': 'DATOS DEL CLIENTE', 'color': '833C0C', 'texto': 'FFFFFF'},
        4: {'titulo': 'DETALLE FINANCIERO', 'color': '5B3151', 'texto': 'FFFFFF'},
        5: {'titulo': 'ADJUNTOS', 'color': '000000', 'texto': 'FFFFFF'}
    }

    def __init__(self, nombre_archivo):
        self.nombre_archivo = nombre_archivo
        self.fila_inicio_datos = 7 

    def generar(self, datos_completos):
        if not datos_completos: return False
        try:
            # Ordenar por número de ítem
            datos_ordenados = sorted(datos_completos, key=lambda x: x.get('Numero', 999))
            
            # Preparar DataFrame con solo las columnas definidas
            claves_columnas = [col[0] for col in self.COLUMNAS_DEF]
            df = pd.DataFrame(datos_ordenados)
            
            # Asegurar que todas las columnas existan aunque no haya datos
            for clave in claves_columnas:
                if clave not in df.columns: df[clave] = ""
            
            # Filtrar y ordenar columnas según definición
            df = df[claves_columnas]
            
            # Escribir Excel básico
            with pd.ExcelWriter(self.nombre_archivo, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, header=False, sheet_name='Reporte DIAN', startrow=self.fila_inicio_datos-1)
            
            # Aplicar estilos
            self._aplicar_diseno_premium(df)
            return True
        except Exception as e:
            log(98, f"❌ Error generando Excel: {e}", "ERROR")
            return False
    
    def _aplicar_diseno_premium(self, df):
        try:
            wb = load_workbook(self.nombre_archivo)
            ws = wb.active
            
            # ESTILOS BASE
            font_dashboard_lbl = Font(name='Segoe UI', size=9, color="666666")
            font_dashboard_val = Font(name='Segoe UI', size=14, bold=True, color="1F4E78")
            font_grupo = Font(name='Segoe UI', size=10, bold=True, color="FFFFFF")
            font_header = Font(name='Segoe UI', size=9, bold=True, color="FFFFFF")
            font_data = Font(name='Segoe UI', size=9)
            
            border_thin = Side(style='thin', color="BFBFBF")
            border_med = Side(style='medium', color="FFFFFF")
            borde_cuadro = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
            
            align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)

            # --- SECCION DASHBOARD (ENCABEZADO SUPERIOR) ---
            ws['B2'] = "REPORTE DETALLADO DE FACTURACIÓN ELECTRÓNICA"
            ws['B2'].font = Font(name='Segoe UI', size=16, bold=True, color="404040")
            
            # KPI: Total Documentos
            ws['B3'] = "TOTAL DOCUMENTOS"
            ws['B3'].font = font_dashboard_lbl
            ws['B3'].alignment = align_center
            ws['B4'] = len(df)
            ws['B4'].font = font_dashboard_val
            ws['B4'].alignment = align_center
            ws['B4'].border = Border(bottom=Side(style='thick', color="1F4E78"))
            
            # KPI: Fecha Generación
            ws['D3'] = "FECHA GENERACIÓN"
            ws['D3'].font = font_dashboard_lbl
            ws['D3'].alignment = align_center
            ws['D4'] = datetime.now().strftime("%d/%m/%Y %H:%M")
            ws['D4'].font = Font(name='Segoe UI', size=11, bold=True, color="404040")
            ws['D4'].alignment = align_center
            ws['D4'].border = Border(bottom=Side(style='thick', color="1F4E78"))

            # --- HEADERS AGRUPADOS (FILA 5) ---
            col_idx = 1
            for grp_id, info in self.GRUPOS_INFO.items():
                cols_grupo = [c for c in self.COLUMNAS_DEF if c[3] == grp_id]
                if not cols_grupo: continue
                
                start_col = col_idx
                end_col = col_idx + len(cols_grupo) - 1
                
                ws.merge_cells(start_row=5, start_column=start_col, end_row=5, end_column=end_col)
                cell = ws.cell(row=5, column=start_col)
                cell.value = info['titulo']
                cell.font = font_grupo
                cell.fill = PatternFill(start_color=info['color'], end_color=info['color'], fill_type="solid")
                cell.alignment = align_center
                cell.border = Border(right=border_med, left=border_med)
                
                col_idx += len(cols_grupo)

            # --- HEADERS DE COLUMNAS (FILA 6) ---
            col_idx = 1
            for def_col in self.COLUMNAS_DEF:
                grp_id = def_col[3]
                color_base = self.GRUPOS_INFO[grp_id]['color']
                
                cell = ws.cell(row=6, column=col_idx)
                cell.value = def_col[1]
                cell.fill = PatternFill(start_color=color_base, end_color=color_base, fill_type="solid")
                cell.font = font_header
                cell.alignment = align_center
                cell.border = borde_cuadro
                ws.column_dimensions[get_column_letter(col_idx)].width = def_col[2]
                col_idx += 1
            
            ws.row_dimensions[5].height = 25
            ws.row_dimensions[6].height = 40

            # --- FORMATO DE DATOS (FILAS 7 en adelante) ---
            cols_moneda = [i+1 for i, c in enumerate(self.COLUMNAS_DEF) if c[3] == 4]
            # Busqueda segura de indices de columnas especiales
            col_pdf = next((i+1 for i, c in enumerate(self.COLUMNAS_DEF) if c[0] == 'Ruta_PDF'), -1)
            col_factura = next((i+1 for i, c in enumerate(self.COLUMNAS_DEF) if c[0] == 'Numero_Factura'), -1)
            col_eventos = next((i+1 for i, c in enumerate(self.COLUMNAS_DEF) if c[0] == 'Eventos'), -1)

            for row in ws.iter_rows(min_row=self.fila_inicio_datos):
                fill = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid") if row[0].row % 2 == 0 else None
                
                for cell in row:
                    if fill: cell.fill = fill
                    cell.font = font_data
                    cell.border = borde_cuadro
                    cell.alignment = Alignment(vertical='center', wrap_text=False)
                    
                    if cell.col_idx in cols_moneda:
                        cell.number_format = '_-$ * #,##0.00_-;-$ * #,##0.00_-;_-;_-@'
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                    
                    if cell.col_idx == col_factura:
                        cell.font = Font(name='Segoe UI', size=9, bold=True)
                        cell.alignment = align_center
                        
                    # Resaltar Eventos si existen
                    if cell.col_idx == col_eventos and cell.value and len(str(cell.value)) > 2:
                        cell.font = Font(name='Segoe UI', size=9, color="E26B0A", bold=True) # Naranja
                        cell.alignment = align_center

                    if cell.col_idx == col_pdf:
                        path = cell.value
                        if path and isinstance(path, str) and os.path.exists(path):
                            cell.value = "Abrir PDF"
                            cell.hyperlink = path
                            cell.font = Font(name='Segoe UI', size=9, color="0000FF", underline="single", bold=True)
                            cell.alignment = align_center
                        elif not path:
                            cell.value = "-"
                            cell.alignment = align_center

            ws.freeze_panes = 'E7' 
            ws.auto_filter.ref = f"A6:{get_column_letter(len(self.COLUMNAS_DEF))}{ws.max_row}"
            wb.save(self.nombre_archivo)
            
        except Exception as e:
            log(98, f"Error formato visual Pro: {e}", "ERROR")

def generar_excel_final(nombre_archivo, datos_completos):
    generador = GeneradorExcel(nombre_archivo)
    return generador.generar(datos_completos)


# ═══════════════════════════════════════════════════════════════════════════════
# HOJA "NO PROCESADOS" — captura todo lo que no llegó al Excel principal
# ═══════════════════════════════════════════════════════════════════════════════

def _motivo_legible(mensaje: str) -> str:
    """Convierte el mensaje técnico del resultado en descripción legible."""
    if not mensaje:
        return 'Error desconocido'
    mapa = [
        ('No existe en DIAN',           'CUFE no registrado en la DIAN'),
        ('Bloqueado anti-robot',         'Bloqueado por controles de seguridad DIAN — requiere intervención humana'),
        ('sin_nit',                      'CUFE sin NIT asociado en el archivo de entrada'),
        ('Campo CUFE no encontrado',     'Error de conexión: campo de búsqueda no disponible'),
        ('Campo NIT no encontrado',      'Error de conexión: campo NIT no disponible en el portal'),
        ('Botón búsqueda no encontrado', 'Error de conexión: botón de búsqueda no disponible'),
        ('Timeout botón PDF',            'Timeout: el botón de descarga PDF no apareció'),
        ('PDF no detectado',             'Timeout: PDF no fue detectado en carpeta de descargas'),
        ('Timeout PDF',                  'Timeout: descarga de PDF excedió el tiempo límite'),
        ('Falló tras',                   'Error definitivo tras múltiples reintentos'),
        ('No procesado (intento',        'Cola de reintentos agotada sin éxito'),
    ]
    for clave, descripcion in mapa:
        if clave in mensaje:
            return descripcion
    if 'Error:' in mensaje:
        detalle = mensaje[mensaje.index('Error:') + 6:].strip()[:80]
        return f'Error técnico: {detalle}'
    return mensaje[:100]


def agregar_hoja_no_procesados(nombre_archivo: str,
                                errores_descarga: list,
                                invalidos_validacion: list = None) -> bool:
    """
    Agrega la hoja 'No Procesados' al Excel ya generado por generar_excel_final().

    Args:
        nombre_archivo      : Ruta del Excel existente (Reporte_YYYYMMDD_HHMMSS.xlsx).
        errores_descarga    : Lista de dicts resultado del orquestador con
                              estado in ('error', 'no_encontrado').
        invalidos_validacion: Lista opcional de dicts {'cufe', 'razon', 'linea'}
                              con CUFEs que fallaron ANTES de intentar descargarse.
                              También acepta strings planos por compatibilidad con ui/app.py.

    Returns:
        True si se generó correctamente, False en caso de error.
    """
    from datetime import datetime

    if not os.path.exists(nombre_archivo):
        log(98, f"❌ No se puede agregar hoja 'No Procesados': archivo no encontrado", "ERROR")
        return False

    # ── Construir filas normalizadas ──────────────────────────────────────────
    filas = []
    num = 1

    # 1. Inválidos de validación (intentos = 0, nunca se intentó la descarga)
    if invalidos_validacion:
        ts_ahora = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        for inv in invalidos_validacion:
            if isinstance(inv, dict):
                cufe  = inv.get('cufe', '')
                razon = inv.get('razon', 'Formato inválido')
                linea = inv.get('linea', '')
            else:                          # string plano (compat. ui/app.py)
                cufe, razon, linea = str(inv), 'Formato de CUFE inválido', ''

            filas.append({
                'Numero':        num,
                'CUFE':          cufe,
                'Motivo':        f'Formato CUFE inválido: {razon}',
                'Intentos':      0,
                'Timestamp':     ts_ahora,
                'Observaciones': f'Línea {linea} del archivo de entrada' if linea else 'Detectado en validación previa',
            })
            num += 1

    # 2. Errores y no-encontrados de la descarga
    for r in errores_descarga:
        mensaje = r.get('mensaje', '')
        filas.append({
            'Numero':        num,
            'CUFE':          r.get('cufe', ''),
            'Motivo':        _motivo_legible(mensaje),
            'Intentos':      r.get('intento', 1),
            'Timestamp':     r.get('timestamp', datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
            'Observaciones': mensaje if len(mensaje) > 5 else '',
        })
        num += 1

    if not filas:
        log(98, "ℹ️  Sin errores que registrar — hoja 'No Procesados' omitida", "INFO")
        return True

    # ── Paleta y estilos (consistentes con GeneradorExcel) ───────────────────
    COLOR_HDR   = '7B1C1C'   # rojo oscuro — tema error
    COLOR_PAR   = 'FFF5F5'   # rosa muy suave para filas alternas
    BLANCO      = 'FFFFFF'

    font_titulo  = Font(name='Segoe UI', size=16, bold=True,  color='404040')
    font_kpi_lbl = Font(name='Segoe UI', size=9,              color='666666')
    font_kpi_val = Font(name='Segoe UI', size=14, bold=True,  color=COLOR_HDR)
    font_grp     = Font(name='Segoe UI', size=10, bold=True,  color=BLANCO)
    font_hdr     = Font(name='Segoe UI', size=9,  bold=True,  color=BLANCO)
    font_data    = Font(name='Segoe UI', size=9)
    font_cufe    = Font(name='Consolas', size=8,              color='444444')
    font_motivo  = Font(name='Segoe UI', size=9,  bold=True,  color='C0392B')

    side_thin = Side(style='thin',   color='BFBFBF')
    side_med  = Side(style='medium', color=BLANCO)
    borde     = Border(left=side_thin, right=side_thin, top=side_thin, bottom=side_thin)

    ac = Alignment(horizontal='center', vertical='center', wrap_text=True)
    al = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    aw = Alignment(horizontal='left',   vertical='center', wrap_text=False)

    # (clave_dict, título_col, ancho)
    COLUMNAS = [
        ('Numero',        'N°',               6),
        ('CUFE',          'CUFE',            36),
        ('Motivo',        'Motivo del Error', 35),
        ('Intentos',      'Intentos',         10),
        ('Timestamp',     'Último Intento',   18),
        ('Observaciones', 'Observaciones',    35),
    ]
    N = len(COLUMNAS)

    try:
        wb = load_workbook(nombre_archivo)

        if 'No Procesados' in wb.sheetnames:
            del wb['No Procesados']

        ws = wb.create_sheet('No Procesados')

        # ── Fila 2: Título ────────────────────────────────────────────────────
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=N)
        c = ws.cell(row=2, column=1)
        c.value     = 'REPORTE DE DOCUMENTOS NO PROCESADOS'
        c.font      = font_titulo
        c.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[2].height = 28

        # ── Filas 3-4: Dashboard KPI ──────────────────────────────────────────
        ws['A3'] = 'TOTAL NO PROCESADOS'
        ws['A3'].font      = font_kpi_lbl
        ws['A3'].alignment = ac

        ws['C3'] = 'FECHA GENERACIÓN'
        ws['C3'].font      = font_kpi_lbl
        ws['C3'].alignment = ac

        ws['A4'] = len(filas)
        ws['A4'].font      = font_kpi_val
        ws['A4'].alignment = ac
        ws['A4'].border    = Border(bottom=Side(style='thick', color=COLOR_HDR))

        ws['C4'] = datetime.now().strftime('%d/%m/%Y %H:%M')
        ws['C4'].font      = Font(name='Segoe UI', size=11, bold=True, color='404040')
        ws['C4'].alignment = ac
        ws['C4'].border    = Border(bottom=Side(style='thick', color=COLOR_HDR))

        ws.row_dimensions[3].height = 16
        ws.row_dimensions[4].height = 24

        # ── Fila 6: Cabecera de grupo ─────────────────────────────────────────
        ws.merge_cells(start_row=6, start_column=1, end_row=6, end_column=N)
        c = ws.cell(row=6, column=1)
        c.value     = 'DETALLE DE ERRORES Y DOCUMENTOS NO PROCESADOS'
        c.font      = font_grp
        c.fill      = PatternFill(start_color=COLOR_HDR, end_color=COLOR_HDR, fill_type='solid')
        c.alignment = ac
        c.border    = Border(left=side_med, right=side_med)
        ws.row_dimensions[6].height = 25

        # ── Fila 7: Encabezados de columna ────────────────────────────────────
        for col_i, (_, titulo, ancho) in enumerate(COLUMNAS, 1):
            c = ws.cell(row=7, column=col_i)
            c.value     = titulo
            c.font      = font_hdr
            c.fill      = PatternFill(start_color=COLOR_HDR, end_color=COLOR_HDR, fill_type='solid')
            c.alignment = ac
            c.border    = borde
            ws.column_dimensions[get_column_letter(col_i)].width = ancho
        ws.row_dimensions[7].height = 35

        # ── Filas 8+: Datos ───────────────────────────────────────────────────
        FILA_DATOS = 8
        for r_idx, fila in enumerate(filas):
            excel_row = FILA_DATOS + r_idx
            fill_par  = PatternFill(start_color=COLOR_PAR, end_color=COLOR_PAR, fill_type='solid') \
                        if excel_row % 2 == 0 else None

            for col_i, (key, titulo, ancho) in enumerate(COLUMNAS, 1):
                c = ws.cell(row=excel_row, column=col_i)
                c.value  = fila.get(key, '')
                c.border = borde
                if fill_par:
                    c.fill = fill_par

                if key == 'Numero':
                    c.font, c.alignment = font_data,   ac
                elif key == 'CUFE':
                    c.font, c.alignment = font_cufe,   aw
                elif key == 'Motivo':
                    c.font, c.alignment = font_motivo, al
                elif key == 'Intentos':
                    c.font, c.alignment = font_data,   ac
                elif key == 'Timestamp':
                    c.font, c.alignment = font_data,   ac
                else:
                    c.font, c.alignment = font_data,   al

        # ── Freeze + Auto-filtro ──────────────────────────────────────────────
        ws.freeze_panes = 'A8'
        ws.auto_filter.ref = f"A7:{get_column_letter(N)}{FILA_DATOS + len(filas) - 1}"

        wb.save(nombre_archivo)
        log(98, f"✅ Hoja 'No Procesados' generada: {len(filas)} registros", "OK")
        return True

    except Exception as e:
        log(98, f"❌ Error generando hoja 'No Procesados': {e}", "ERROR")
        return False