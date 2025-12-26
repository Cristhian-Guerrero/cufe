import time
import os
import json
import threading
import queue
import re
import hashlib
import atexit
from datetime import datetime
import pandas as pd
import pdfplumber
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from DrissionPage import ChromiumPage, ChromiumOptions

# === CONFIGURACIÓN ===
DIAN_URL = "https://catalogo-vpfe.dian.gov.co/User/SearchDocument"
CARPETA_PDFS = "facturas_pdfs_descargados"
ARCHIVO_MAPPING = "mapping_cufes_pdfs.json"
ARCHIVO_EXCEL = f"Facturas_Completas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
NUM_NAVEGADORES = 10  # 🚀 10 navegadores en paralelo
MAX_REINTENTOS = 2    # Reintentar 2 veces si falla

# === COLAS DE PIPELINE ===
cola_trabajo = queue.Queue()
cola_reintentos = queue.Queue()  # Cola para reintentos
cola_pdfs_procesar = queue.Queue()
cola_resultados = queue.Queue()

# === LOCKS ===
lock_consola = threading.Lock()
lock_mapping = threading.Lock()
lock_excel = threading.Lock()
lock_reintentos = threading.Lock()

# === ESTADO GLOBAL ===
mapping_cufes = {}
navegadores_activos = []
orden_original = []
datos_completos = []
intentos_por_cufe = {}  # Tracking de intentos

def log(nav_id, mensaje, nivel="INFO"):
    """Log con colores"""
    colores = {
        "INFO": "\033[94m",
        "OK": "\033[92m",
        "WARN": "\033[93m",
        "ERROR": "\033[91m",
        "DEBUG": "\033[96m",
        "CRIT": "\033[95m",
        "EXCEL": "\033[93m",
        "RETRY": "\033[95m"
    }
    
    with lock_consola:
        ts = datetime.now().strftime("%H:%M:%S.%f")[:-3]
        if nav_id == 0:
            prefix = "[SYS]"
        elif nav_id == 99:
            prefix = "[EXT]"
        elif nav_id == 98:
            prefix = "[XLS]"
        else:
            prefix = f"[N{nav_id}]"
        
        color = colores.get(nivel, "")
        print(f"{color}{ts} {prefix} [{nivel}] {mensaje}\033[0m")

def limpiar_al_salir():
    """Limpia navegadores"""
    log(0, "🧹 Limpiando...", "WARN")
    for page in navegadores_activos:
        try:
            page.quit()
        except:
            pass

atexit.register(limpiar_al_salir)

def guardar_mapping():
    """Guarda mapping JSON"""
    with lock_mapping:
        try:
            with open(ARCHIVO_MAPPING, 'w', encoding='utf-8') as f:
                json.dump(mapping_cufes, f, indent=2, ensure_ascii=False)
        except:
            pass

# === EXTRACCIÓN DE DATOS PDF ===

def limpiar_texto(texto):
    if not texto:
        return ""
    return re.sub(r'\s+', ' ', texto).strip()

def limpiar_nit(texto):
    if not texto:
        return ""
    match = re.search(r'([\d\.-]+)', texto)
    return match.group(1) if match else texto

def extraer_datos_pdf(ruta_pdf, cufe_original, numero):
    """Extrae todos los datos del PDF"""
    datos = {
        'Numero': numero,
        'CUFE': cufe_original,
        'Numero_Factura': '',
        'Prefijo': '',
        'Folio': '',
        'Fecha_Emision': '',
        'Fecha_Vencimiento': '',
        'Forma_Pago': '',
        'Medio_Pago': '',
        'Emisor_RazonSocial': '',
        'Emisor_NIT': '',
        'Emisor_Direccion': '',
        'Emisor_Ciudad': '',
        'Emisor_Departamento': '',
        'Emisor_Telefono': '',
        'Emisor_Email': '',
        'Receptor_RazonSocial': '',
        'Receptor_NIT': '',
        'Receptor_Direccion': '',
        'Receptor_Ciudad': '',
        'Receptor_Departamento': '',
        'Receptor_Email': '',
        'Subtotal': '',
        'IVA': '',
        'Total_Factura': '',
        'Numero_Autorizacion': '',
        'Ruta_PDF': os.path.abspath(ruta_pdf),
        'Notas': '',
        'Estado': '✅ Procesado'
    }
    
    try:
        with pdfplumber.open(ruta_pdf) as pdf:
            texto_completo = ""
            for pagina in pdf.pages:
                texto_completo += pagina.extract_text() + "\n"
            
            # CUFE
            m = re.search(r'CUFE\s*:?\s*([\w\n]+)', texto_completo)
            if m:
                datos['CUFE'] = m.group(1).replace('\n', '').strip()
            
            # Número factura
            m = re.search(r'Número de Factura:\s*([A-Z0-9\-]+)', texto_completo)
            if m:
                numero_factura = m.group(1)
                datos['Numero_Factura'] = numero_factura
                partes = numero_factura.split('-')
                if len(partes) == 2:
                    datos['Prefijo'] = partes[0]
                    datos['Folio'] = partes[1]
            
            # Fechas
            m = re.search(r'Fecha de Emisión:\s*(\d{2}/\d{2}/\d{4})', texto_completo)
            if m:
                datos['Fecha_Emision'] = m.group(1)
            
            m = re.search(r'Fecha de Vencimiento:\s*(\d{2}/\d{2}/\d{4})', texto_completo)
            if m:
                datos['Fecha_Vencimiento'] = m.group(1)
            
            # Bloques
            bloque_emisor = re.search(
                r'Datos del Emisor.*?Datos del Adquiriente',
                texto_completo,
                re.DOTALL | re.IGNORECASE
            )
            txt_emisor = bloque_emisor.group(0) if bloque_emisor else texto_completo
            
            bloque_receptor = re.search(
                r'Datos del Adquiriente.*?(?:Detalles de Productos|Observaciones)',
                texto_completo,
                re.DOTALL | re.IGNORECASE
            )
            txt_receptor = bloque_receptor.group(0) if bloque_receptor else ""
            
            # Emisor
            m = re.search(r'Razón Social:\s*([^\n]+)', txt_emisor)
            if m:
                datos['Emisor_RazonSocial'] = limpiar_texto(m.group(1))
            
            m = re.search(r'Nit del Emisor:\s*([^\n]+)', txt_emisor)
            if m:
                datos['Emisor_NIT'] = limpiar_nit(m.group(1))
            
            m = re.search(r'Dirección:\s*([^\n]+)', txt_emisor)
            if m:
                datos['Emisor_Direccion'] = limpiar_texto(m.group(1))
            
            m = re.search(r'Ciudad:\s*([^\n]+)', txt_emisor)
            if m:
                datos['Emisor_Ciudad'] = limpiar_texto(m.group(1))
            
            m = re.search(r'Departamento:\s*([^\n]+)', txt_emisor)
            if m:
                datos['Emisor_Departamento'] = limpiar_texto(m.group(1))
            
            m = re.search(r'(?:Teléfono|Telefono):\s*([^\n]+)', txt_emisor)
            if m:
                datos['Emisor_Telefono'] = limpiar_texto(m.group(1))
            
            m = re.search(r'Email:\s*([^\n]+)', txt_emisor)
            if m:
                datos['Emisor_Email'] = limpiar_texto(m.group(1))
            
            # Receptor
            if txt_receptor:
                m = re.search(r'(?:Nombre o )?Razón Social:\s*([^\n]+)', txt_receptor)
                if m:
                    datos['Receptor_RazonSocial'] = limpiar_texto(m.group(1))
                
                m = re.search(r'Número Documento:\s*([^\n]+)', txt_receptor)
                if m:
                    datos['Receptor_NIT'] = limpiar_nit(m.group(1))
                
                m = re.search(r'Dirección:\s*([^\n]+)', txt_receptor)
                if m:
                    datos['Receptor_Direccion'] = limpiar_texto(m.group(1))
                
                m = re.search(r'Ciudad:\s*([^\n]+)', txt_receptor)
                if m:
                    datos['Receptor_Ciudad'] = limpiar_texto(m.group(1))
                
                m = re.search(r'Departamento:\s*([^\n]+)', txt_receptor)
                if m:
                    datos['Receptor_Departamento'] = limpiar_texto(m.group(1))
                
                m = re.search(r'Email:\s*([^\n]+)', txt_receptor)
                if m:
                    datos['Receptor_Email'] = limpiar_texto(m.group(1))
            
            # Financiero
            m = re.search(r'Total factura.*?(?:COP\s+)?\$?\s*([\d\.,]+)', texto_completo)
            if m:
                datos['Total_Factura'] = m.group(1)
            
            m = re.search(r'Subtotal\s*[\n\r]*\s*(?:COP)?\s*\$?\s*([\d\.,]+)', texto_completo)
            if m:
                datos['Subtotal'] = m.group(1)
            
            m = re.search(r'Total impuesto.*?(=).*?([\d\.,]+)', texto_completo, re.DOTALL)
            if not m:
                m = re.search(r'(?:^|\n)IVA\s+([\d,\.]+)', texto_completo)
            if m:
                datos['IVA'] = m.group(2) if len(m.groups()) > 1 else m.group(1)
            
            # Pagos
            m = re.search(r'Forma de pago:\s*([^\n]+)', texto_completo)
            if m:
                datos['Forma_Pago'] = limpiar_texto(m.group(1))
            
            m = re.search(r'Medio de Pago:\s*([^\n]+)', texto_completo)
            if m:
                datos['Medio_Pago'] = limpiar_texto(m.group(1))
            
            # Legal
            m = re.search(r'Numero de Autorización:\s*(\d+)', texto_completo)
            if m:
                datos['Numero_Autorizacion'] = m.group(1)
    
    except Exception as e:
        log(99, f"Error extrayendo: {str(e)[:40]}", "ERROR")
        datos['Estado'] = '❌ Error'
    
    return datos

# === CLOUDFLARE BYPASS ===

class CloudflareBypass:
    def __init__(self, page, nav_id):
        self.page = page
        self.nav_id = nav_id
    
    def intentar(self, timeout=8, max_intentos=2):
        for intento in range(max_intentos):
            try:
                iframe = self.page.ele('css:iframe[src*="cloudflare"]', timeout=timeout)
                if not iframe:
                    return True
                
                body = iframe.ele('tag:body', timeout=3)
                if not body or not body.shadow_root:
                    time.sleep(1)
                    continue
                
                checkbox = body.shadow_root.ele('tag:input', timeout=2)
                if not checkbox:
                    time.sleep(1)
                    continue
                
                if checkbox.states.is_checked:
                    return True
                
                checkbox.click()
                time.sleep(4)
                
                if checkbox.states.is_checked:
                    log(self.nav_id, "✓ Validado", "DEBUG")
                    return True
                
            except:
                if intento < max_intentos - 1:
                    time.sleep(2)
        
        return False

# === DESCARGA ===

def generar_nombre_unico(cufe, nav_id):
    timestamp_micro = int(time.time() * 1000000)
    hash_parte = hashlib.md5(f"{cufe}_{nav_id}_{timestamp_micro}".encode()).hexdigest()[:12]
    return f"FACTURA_{cufe[:20]}_{hash_parte}.pdf"

def inicializar_navegador(nav_id):
    log(nav_id, "🌐 Iniciando...", "INFO")
    
    port = 9700 + (nav_id * 5)
    user_data = os.path.join(os.getcwd(), f".chrome_dian_{nav_id}")
    
    co = ChromiumOptions()
    co.set_local_port(port)
    co.headless(False)
    co.set_argument('--no-sandbox')
    co.set_argument('--disable-dev-shm-usage')
    co.set_argument('--disable-gpu')
    co.set_argument('--window-size=1000,700')
    co.set_argument(f'--user-data-dir={user_data}')
    co.set_argument('--disable-blink-features=AutomationControlled')
    
    # Posición en cuadrícula 2x5
    fila = (nav_id - 1) // 5
    columna = (nav_id - 1) % 5
    x = columna * 300
    y = fila * 400
    co.set_argument(f'--window-position={x},{y}')
    
    os.makedirs(CARPETA_PDFS, exist_ok=True)
    ruta_absoluta = os.path.abspath(CARPETA_PDFS)
    co.set_download_path(ruta_absoluta)
    
    # Forzar descarga
    co.set_pref('download.default_directory', ruta_absoluta)
    co.set_pref('download.prompt_for_download', False)
    co.set_pref('plugins.always_open_pdf_externally', True)
    
    try:
        page = ChromiumPage(addr_or_opts=co)
        page.set.timeouts(20)
        page.set.download_path(ruta_absoluta)
        navegadores_activos.append(page)
        log(nav_id, f"✓ OK (:{port})", "OK")
        return page
    except Exception as e:
        log(nav_id, f"❌ Error: {e}", "ERROR")
        return None

def detectar_pdf(cufe, nav_id, archivos_antes, timeout=20):
    """Detecta PDF nuevo"""
    log(nav_id, f"⏳ Esperando PDF ({timeout}s)...", "INFO")
    
    tiempo_inicio = time.time()
    cufe_parcial = cufe[:20]
    
    while (time.time() - tiempo_inicio) < timeout:
        time.sleep(0.5)
        
        try:
            archivos_ahora = set(os.listdir(CARPETA_PDFS))
            archivos_nuevos = archivos_ahora - archivos_antes
            
            pdfs_cufe = [
                f for f in archivos_nuevos 
                if f.endswith('.pdf') and cufe_parcial in f and not f.endswith('.crdownload')
            ]
            
            if pdfs_cufe:
                pdf_nombre = pdfs_cufe[0]
                ruta_pdf = os.path.join(CARPETA_PDFS, pdf_nombre)
                tamanio = os.path.getsize(ruta_pdf)
                
                if tamanio > 1000:
                    nombre_nuevo = generar_nombre_unico(cufe, nav_id)
                    ruta_nueva = os.path.join(CARPETA_PDFS, nombre_nuevo)
                    
                    try:
                        os.rename(ruta_pdf, ruta_nueva)
                        log(nav_id, f"✓ {nombre_nuevo}", "OK")
                        
                        with lock_mapping:
                            mapping_cufes[cufe] = nombre_nuevo
                        
                        return ruta_nueva
                    except:
                        return ruta_pdf
                        
        except:
            continue
    
    log(nav_id, f"❌ TIMEOUT", "ERROR")
    return None

def descargar_cufe(page, bypass, cufe, numero, total, nav_id, intento=1):
    if intento > 1:
        log(nav_id, "="*50, "RETRY")
        log(nav_id, f"🔄 REINTENTO {intento}/{MAX_REINTENTOS}", "RETRY")
        log(nav_id, f"📥 CUFE {numero}/{total}", "RETRY")
        log(nav_id, "="*50, "RETRY")
    else:
        log(nav_id, "="*50, "INFO")
        log(nav_id, f"📥 CUFE {numero}/{total}", "INFO")
        log(nav_id, "="*50, "INFO")
    
    resultado = {
        'numero': numero,
        'cufe': cufe,
        'estado': 'error',
        'pdf': None,
        'ruta_pdf': None,
        'mensaje': '',
        'intento': intento
    }
    
    archivos_antes = set(os.listdir(CARPETA_PDFS))
    
    try:
        if "SearchDocument" not in page.url:
            page.get(DIAN_URL)
            time.sleep(2)
            bypass.intentar()
        
        campo_cufe = page.ele('#DocumentKey', timeout=8)
        if not campo_cufe:
            resultado['mensaje'] = "Campo CUFE no encontrado"
            return resultado
        
        log(nav_id, "⌨️ Ingresando...", "INFO")
        campo_cufe.clear()
        time.sleep(0.5)
        campo_cufe.input(cufe, clear=True)
        time.sleep(1)
        
        bypass.intentar(timeout=10)
        time.sleep(2)
        
        boton_buscar = page.ele('css:button.search-document', timeout=8)
        if not boton_buscar:
            resultado['mensaje'] = "Botón búsqueda no encontrado"
            return resultado
        
        log(nav_id, "🔍 Buscando...", "INFO")
        boton_buscar.click()
        time.sleep(4)
        
        if page.ele('text:Documento no encontrado', timeout=2):
            log(nav_id, "⚠️ NO ENCONTRADO", "WARN")
            resultado['estado'] = 'no_encontrado'
            resultado['mensaje'] = "No existe en DIAN"
            return resultado
        
        bypass.intentar(timeout=10)
        time.sleep(2)
        
        log(nav_id, "🔎 Buscando PDF...", "INFO")
        
        boton_pdf = None
        tiempo_busqueda = time.time()
        timeout_pdf = 45
        
        while not boton_pdf and (time.time() - tiempo_busqueda) < timeout_pdf:
            try:
                botones = page.eles('tag:a', timeout=2)
                for boton in botones:
                    texto = boton.text.lower()
                    if ("descargar" in texto and "pdf" in texto) or "descargar pdf" in texto:
                        boton_pdf = boton
                        log(nav_id, f"✓ Botón encontrado", "OK")
                        break
                
                if not boton_pdf:
                    time.sleep(2)
                
            except:
                time.sleep(2)
        
        if not boton_pdf:
            log(nav_id, "❌ Botón PDF no apareció", "ERROR")
            resultado['mensaje'] = "Timeout botón PDF"
            return resultado
        
        log(nav_id, "📥 Descargando...", "INFO")
        bypass.intentar(timeout=5)
        time.sleep(2)
        
        boton_pdf.click()
        log(nav_id, "✓ Click OK", "OK")
        time.sleep(1)
        
        ruta_pdf = detectar_pdf(cufe, nav_id, archivos_antes, timeout=20)
        
        if ruta_pdf:
            log(nav_id, "✅ EXITOSO", "OK")
            resultado['estado'] = 'exitoso'
            resultado['pdf'] = os.path.basename(ruta_pdf)
            resultado['ruta_pdf'] = ruta_pdf
            resultado['mensaje'] = "OK"
        else:
            log(nav_id, "❌ PDF no detectado", "ERROR")
            resultado['mensaje'] = "Timeout PDF"
        
    except Exception as e:
        log(nav_id, f"❌ Error: {str(e)[:50]}", "ERROR")
        resultado['mensaje'] = f"Error: {str(e)[:80]}"
    
    return resultado

def trabajador_descarga(nav_id):
    """Worker con reintentos automáticos"""
    log(nav_id, f"🚀 Iniciando", "INFO")
    
    page = inicializar_navegador(nav_id)
    if not page:
        return
    
    bypass = CloudflareBypass(page, nav_id)
    
    log(nav_id, "🌐 Primera navegación...", "INFO")
    page.get(DIAN_URL)
    time.sleep(3)
    bypass.intentar()
    
    while True:
        try:
            item = cola_trabajo.get(timeout=5)
            
            if item is None:
                log(nav_id, "🏁 Fin", "INFO")
                break
            
            cufe, numero, total = item
            
            # Primer intento
            resultado = descargar_cufe(page, bypass, cufe, numero, total, nav_id, intento=1)
            
            # Si falló y puede reintentar
            if resultado['estado'] in ['error'] and resultado['mensaje'] != "No existe en DIAN":
                with lock_reintentos:
                    if cufe not in intentos_por_cufe:
                        intentos_por_cufe[cufe] = 1
                    else:
                        intentos_por_cufe[cufe] += 1
                    
                    intentos_actuales = intentos_por_cufe[cufe]
                
                if intentos_actuales < MAX_REINTENTOS:
                    log(nav_id, f"🔄 Programando reintento ({intentos_actuales+1}/{MAX_REINTENTOS})", "RETRY")
                    cola_reintentos.put((cufe, numero, total))
                else:
                    log(nav_id, f"⚠️ Máximo de reintentos alcanzado", "WARN")
                    cola_resultados.put(resultado)
            else:
                cola_resultados.put(resultado)
            
            # Si fue exitoso, enviar a extractor
            if resultado['estado'] == 'exitoso' and resultado['ruta_pdf']:
                cola_pdfs_procesar.put({
                    'numero': numero,
                    'cufe': cufe,
                    'ruta_pdf': resultado['ruta_pdf']
                })
                log(nav_id, f"→ Enviado a extractor", "DEBUG")
            
            time.sleep(2)
            
        except queue.Empty:
            break
        except Exception as e:
            log(nav_id, f"Error worker: {e}", "ERROR")
            break
    
    try:
        page.quit()
        if page in navegadores_activos:
            navegadores_activos.remove(page)
        log(nav_id, "Navegador cerrado", "INFO")
    except:
        pass

def procesador_reintentos(nav_id):
    """Procesa reintentos en un navegador dedicado"""
    log(nav_id, "🔄 Procesador de reintentos iniciado", "RETRY")
    
    page = inicializar_navegador(nav_id)
    if not page:
        return
    
    bypass = CloudflareBypass(page, nav_id)
    page.get(DIAN_URL)
    time.sleep(3)
    bypass.intentar()
    
    procesados = 0
    
    while True:
        try:
            item = cola_reintentos.get(timeout=10)
            
            if item is None:
                log(nav_id, "🏁 Fin reintentos", "RETRY")
                break
            
            cufe, numero, total = item
            
            with lock_reintentos:
                intento_actual = intentos_por_cufe.get(cufe, 1) + 1
                intentos_por_cufe[cufe] = intento_actual
            
            log(nav_id, f"🔄 Reintentando CUFE #{numero} (intento {intento_actual})", "RETRY")
            
            resultado = descargar_cufe(page, bypass, cufe, numero, total, nav_id, intento=intento_actual)
            
            # Si falló de nuevo y puede reintentar
            if resultado['estado'] == 'error' and intento_actual < MAX_REINTENTOS:
                log(nav_id, f"⚠️ Falló de nuevo, reintentando...", "RETRY")
                cola_reintentos.put((cufe, numero, total))
            else:
                # Ya sea éxito o fallo definitivo
                cola_resultados.put(resultado)
                
                if resultado['estado'] == 'exitoso' and resultado['ruta_pdf']:
                    cola_pdfs_procesar.put({
                        'numero': numero,
                        'cufe': cufe,
                        'ruta_pdf': resultado['ruta_pdf']
                    })
                    log(nav_id, f"✅ REINTENTO EXITOSO", "OK")
            
            procesados += 1
            time.sleep(3)
            
        except queue.Empty:
            # Si no hay más reintentos, salir
            if procesados > 0:
                log(nav_id, f"✓ Procesados {procesados} reintentos", "RETRY")
            break
        except Exception as e:
            log(nav_id, f"Error reintentos: {e}", "ERROR")
            break
    
    try:
        page.quit()
        log(nav_id, "Navegador de reintentos cerrado", "RETRY")
    except:
        pass

# === EXTRACTOR ===

def trabajador_extractor():
    log(99, "🔍 Extractor iniciado", "OK")
    procesados = 0
    
    while True:
        try:
            item = cola_pdfs_procesar.get(timeout=5)
            
            if item is None:
                log(99, "🏁 Fin extractor", "INFO")
                break
            
            numero = item['numero']
            cufe = item['cufe']
            ruta_pdf = item['ruta_pdf']
            
            log(99, f"📄 Extrayendo #{numero}...", "INFO")
            
            datos = extraer_datos_pdf(ruta_pdf, cufe, numero)
            
            with lock_excel:
                datos_completos.append(datos)
            
            procesados += 1
            log(99, f"✓ Procesado #{numero} ({procesados} total)", "OK")
            
        except queue.Empty:
            time.sleep(1)
        except Exception as e:
            log(99, f"Error: {e}", "ERROR")

# === EXCEL PROFESIONAL ===

def generar_excel_final():
    log(98, "📊 Generando Excel profesional...", "EXCEL")
    
    with lock_excel:
        if not datos_completos:
            log(98, "❌ No hay datos", "ERROR")
            return
        
        datos_ordenados = sorted(datos_completos, key=lambda x: x.get('Numero', 999))
        
        df = pd.DataFrame(datos_ordenados)
        
        columnas_orden = [
            'Numero', 'Estado', 'Numero_Factura', 'Prefijo', 'Folio',
            'Fecha_Emision', 'Fecha_Vencimiento',
            'Emisor_RazonSocial', 'Emisor_NIT', 'Emisor_Ciudad', 'Emisor_Departamento',
            'Emisor_Direccion', 'Emisor_Telefono', 'Emisor_Email',
            'Receptor_RazonSocial', 'Receptor_NIT', 'Receptor_Ciudad', 'Receptor_Departamento',
            'Receptor_Direccion', 'Receptor_Email',
            'Subtotal', 'IVA', 'Total_Factura',
            'Forma_Pago', 'Medio_Pago', 'Numero_Autorizacion',
            'CUFE', 'Ruta_PDF', 'Notas'
        ]
        
        df = df[columnas_orden]
        df.to_excel(ARCHIVO_EXCEL, index=False, sheet_name='Facturas')
        
        log(98, "✓ Datos guardados", "OK")
        aplicar_formato_profesional()
        log(98, "✅ Excel profesional generado", "OK")

def aplicar_formato_profesional():
    try:
        wb = load_workbook(ARCHIVO_EXCEL)
        ws = wb.active
        
        fill_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        font_header = Font(bold=True, color="FFFFFF", size=12, name='Calibri')
        align_header = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        borde_grueso = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        fill_par = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        fill_impar = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        align_centro = Alignment(horizontal='center', vertical='center')
        align_izquierda = Alignment(horizontal='left', vertical='center')
        align_derecha = Alignment(horizontal='right', vertical='center')
        
        for cell in ws[1]:
            cell.fill = fill_header
            cell.font = font_header
            cell.alignment = align_header
            cell.border = borde_grueso
        
        anchos_personalizados = {
            'A': 8, 'B': 12, 'C': 18, 'D': 10, 'E': 10,
            'F': 14, 'G': 14, 'H': 35, 'I': 16, 'J': 20,
            'K': 18, 'L': 35, 'M': 16, 'N': 30, 'O': 35,
            'P': 16, 'Q': 20, 'R': 18, 'S': 35, 'T': 30,
            'U': 15, 'V': 15, 'W': 15, 'X': 18, 'Y': 18,
            'Z': 18, 'AA': 70, 'AB': 15, 'AC': 50
        }
        
        for col, ancho in anchos_personalizados.items():
            ws.column_dimensions[col].width = ancho
        
        col_estado = None
        col_pdf = None
        col_subtotal = None
        col_iva = None
        col_total = None
        
        for idx, cell in enumerate(ws[1], 1):
            if cell.value == 'Estado':
                col_estado = idx
            elif cell.value == 'Ruta_PDF':
                col_pdf = idx
            elif cell.value == 'Subtotal':
                col_subtotal = idx
            elif cell.value == 'IVA':
                col_iva = idx
            elif cell.value == 'Total_Factura':
                col_total = idx
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            fill = fill_par if row_idx % 2 == 0 else fill_impar
            
            for cell_idx, cell in enumerate(row, 1):
                cell.border = borde_grueso
                cell.fill = fill
                
                if cell_idx == 1:
                    cell.alignment = align_centro
                    cell.font = Font(bold=True, size=11)
                elif cell_idx == col_estado:
                    cell.alignment = align_centro
                    cell.font = Font(size=11)
                elif cell_idx in [col_subtotal, col_iva, col_total]:
                    cell.alignment = align_derecha
                    cell.number_format = '$#,##0.00'
                    cell.font = Font(bold=True if cell_idx == col_total else False, size=11)
                else:
                    cell.alignment = align_izquierda
                
                if cell_idx == col_pdf:
                    for dato in datos_completos:
                        if dato.get('Numero') == row_idx - 1:
                            ruta_real = dato.get('Ruta_PDF', '')
                            if ruta_real and os.path.exists(ruta_real):
                                cell.hyperlink = f"file:///{ruta_real.replace(os.sep, '/')}"
                                cell.value = "📄 Ver PDF"
                                cell.font = Font(color="0000FF", underline="single", size=11)
                                cell.alignment = align_centro
                            break
        
        ws.freeze_panes = 'A2'
        ws.row_dimensions[1].height = 30
        
        for row_idx in range(2, ws.max_row + 1):
            ws.row_dimensions[row_idx].height = 20
        
        ws.auto_filter.ref = ws.dimensions
        
        wb.save(ARCHIVO_EXCEL)
        log(98, "✓ Formato aplicado", "OK")
        
    except Exception as e:
        log(98, f"Error formato: {e}", "ERROR")

# === MAIN ===

def main():
    print("\n" + "="*70)
    print("🚀 SISTEMA ULTRA OPTIMIZADO - 10 NAVEGADORES + REINTENTOS")
    print("="*70)
    print()
    
    if not os.path.exists('cufes_test.txt'):
        log(0, "❌ No existe cufes_test.txt", "ERROR")
        return
    
    with open('cufes_test.txt', 'r', encoding='utf-8') as f:
        cufes = [l.strip() for l in f if l.strip()]
    
    global orden_original
    orden_original = cufes.copy()
    
    log(0, f"📋 {len(cufes)} CUFEs", "INFO")
    log(0, f"🚀 {NUM_NAVEGADORES} navegadores paralelos", "INFO")
    log(0, f"🔄 {MAX_REINTENTOS} reintentos automáticos", "INFO")
    log(0, f"📁 {CARPETA_PDFS}/", "INFO")
    log(0, f"📊 {ARCHIVO_EXCEL}", "INFO")
    print()
    
    # Llenar cola
    for i, cufe in enumerate(cufes, 1):
        cola_trabajo.put((cufe, i, len(cufes)))
    
    for _ in range(NUM_NAVEGADORES):
        cola_trabajo.put(None)
    
    # Crear hilos
    threads = []
    
    # 10 navegadores principales
    for i in range(1, NUM_NAVEGADORES + 1):
        t = threading.Thread(target=trabajador_descarga, args=(i,))
        threads.append(t)
    
    # 1 navegador para reintentos
    t_reintentos = threading.Thread(target=procesador_reintentos, args=(99,))
    threads.append(t_reintentos)
    
    # Extractor
    t_extractor = threading.Thread(target=trabajador_extractor)
    threads.append(t_extractor)
    
    tiempo_inicio = time.time()
    
    log(0, "🎬 Iniciando...", "OK")
    
    # Iniciar todos
    for t in threads:
        t.start()
    
    # Esperar navegadores principales
    for t in threads[:NUM_NAVEGADORES]:
        t.join()
    
    log(0, "✓ Descargas principales completadas", "OK")
    
    # Señal de parada para reintentos
    cola_reintentos.put(None)
    t_reintentos.join()
    
    log(0, "✓ Reintentos completados", "OK")
    
    # Parar extractor
    cola_pdfs_procesar.put(None)
    t_extractor.join()
    
    log(0, "✓ Extracción completada", "OK")
    
    # Generar Excel
    generar_excel_final()
    
    duracion = time.time() - tiempo_inicio
    
    guardar_mapping()
    
    # Resultados
    resultados = []
    while not cola_resultados.empty():
        resultados.append(cola_resultados.get())
    
    resultados.sort(key=lambda x: x['numero'])
    
    print("\n" + "="*70)
    print("📊 RESULTADOS FINALES")
    print("="*70)
    
    exitosos = [r for r in resultados if r['estado'] == 'exitoso']
    no_encontrados = [r for r in resultados if r['estado'] == 'no_encontrado']
    errores = [r for r in resultados if r['estado'] == 'error']
    
    log(0, f"✅ Exitosos: {len(exitosos)}/{len(cufes)}", "OK")
    log(0, f"⚠️ No encontrados: {len(no_encontrados)}", "WARN")
    log(0, f"❌ Errores: {len(errores)}", "ERROR")
    log(0, f"⏱️ Tiempo: {duracion:.1f}s ({duracion/60:.1f}min)", "INFO")
    log(0, f"📊 Excel: {ARCHIVO_EXCEL}", "EXCEL")
    log(0, f"📂 PDFs: {CARPETA_PDFS}/", "INFO")
    log(0, f"✨ {len(datos_completos)} registros en Excel", "OK")
    
    # Estadísticas de reintentos
    total_reintentos = sum(1 for r in resultados if r.get('intento', 1) > 1)
    if total_reintentos > 0:
        log(0, f"🔄 {total_reintentos} CUFEs necesitaron reintentos", "RETRY")
    
    # Proyección
    if exitosos:
        promedio = duracion / len(cufes)
        estimacion_100 = (100 * promedio) / NUM_NAVEGADORES / 60
        log(0, f"📈 Estimación 100 CUFEs: ~{estimacion_100:.1f} min", "INFO")
    
    if errores:
        print("\n❌ ERRORES DEFINITIVOS:")
        for r in errores:
            print(f"  CUFE #{r['numero']}: {r['mensaje']} (intentos: {r.get('intento', 1)})")
    
    print("\n✅ Proceso completado\n")

if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        log(0, "\n⚠️ Interrumpido", "WARN")
        limpiar_al_salir()
    except Exception as e:
        log(0, f"\n❌ Error: {e}", "ERROR")
        limpiar_al_salir()