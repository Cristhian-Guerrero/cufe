#!/usr/bin/env python3
"""
Script de prueba para validador v3.1.0
Prueba archivos .txt y .xlsx
"""

import sys
import os

# Simular módulo utils (para testing independiente)
class MockUtils:
    @staticmethod
    def log(nav_id, mensaje, nivel="INFO"):
        colores = {
            "INFO": "\033[94m",
            "OK": "\033[92m",
            "WARN": "\033[93m",
            "ERROR": "\033[91m",
            "CRIT": "\033[95m"
        }
        color = colores.get(nivel, "\033[0m")
        reset = "\033[0m"
        print(f"{color}[{nivel}] {mensaje}{reset}")

sys.modules['utils'] = MockUtils()
from validador_v3 import cargar_cufes

def crear_archivo_txt_prueba():
    """Crea archivo .txt de prueba"""
    contenido = """8d528789cb2547d090fcae2acccfc04e8588a8b80ea7de7bda94ff33c11a0ccf564fa6c73c49ca8ca43809557eaa3a3a
8d528789cb2547d090fcae2acccfc04e8588a8b80ea7de7bda94ff33c11a0ccf564fa6c73c49ca8ca43809557eaa3a3a
INVALIDO_CORTO
gggggggggggggggggggggggggggggggggggggggggggggggggggggggggggggggggggggggggggggggggggggggggggggg
c64601b7a5b65b68d2ff2a7e7c31d48beb674bc0974172368193825144d5aa5c1d7d444412ae5fbc7819e3e70341bc71
"""
    
    with open('/tmp/test_cufes.txt', 'w') as f:
        f.write(contenido)
    
    print("✅ Archivo .txt creado: /tmp/test_cufes.txt")

def crear_archivo_excel_prueba():
    """Crea archivo .xlsx de prueba"""
    try:
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        ws.title = "CUFEs"
        
        # Encabezado
        ws['A1'] = "CUFE"
        
        # CUFEs
        ws['A2'] = "8d528789cb2547d090fcae2acccfc04e8588a8b80ea7de7bda94ff33c11a0ccf564fa6c73c49ca8ca43809557eaa3a3a"
        ws['A3'] = "c64601b7a5b65b68d2ff2a7e7c31d48beb674bc0974172368193825144d5aa5c1d7d444412ae5fbc7819e3e70341bc71"
        ws['A4'] = "INVALIDO_CORTO"
        ws['A5'] = "8d528789cb2547d090fcae2acccfc04e8588a8b80ea7de7bda94ff33c11a0ccf564fa6c73c49ca8ca43809557eaa3a3a"  # Duplicado
        
        wb.save('/tmp/test_cufes.xlsx')
        print("✅ Archivo .xlsx creado: /tmp/test_cufes.xlsx")
        return True
        
    except ImportError:
        print("⚠️  openpyxl no disponible, saltando test de Excel")
        return False

def main():
    print("\n" + "="*70)
    print("🧪 TEST VALIDADOR v3.1.0 - SOPORTE .TXT Y .XLSX")
    print("="*70 + "\n")
    
    # Test 1: Archivo .txt
    print("🔹 TEST 1: Archivo .txt")
    print("-" * 70)
    crear_archivo_txt_prueba()
    
    cufes_txt = cargar_cufes('/tmp/test_cufes.txt')
    
    print(f"\n✅ Resultado .txt: {len(cufes_txt)} CUFEs válidos")
    print(f"   Esperado: 2 CUFEs válidos (1 duplicado, 2 inválidos)")
    
    if len(cufes_txt) == 2:
        print("   ✅ TEST PASADO")
    else:
        print(f"   ❌ TEST FALLIDO (obtenidos: {len(cufes_txt)})")
    
    # Test 2: Archivo .xlsx
    print("\n🔹 TEST 2: Archivo .xlsx")
    print("-" * 70)
    
    if crear_archivo_excel_prueba():
        cufes_xlsx = cargar_cufes('/tmp/test_cufes.xlsx')
        
        print(f"\n✅ Resultado .xlsx: {len(cufes_xlsx)} CUFEs válidos")
        print(f"   Esperado: 2 CUFEs válidos (1 duplicado, 1 inválido)")
        
        if len(cufes_xlsx) == 2:
            print("   ✅ TEST PASADO")
        else:
            print(f"   ❌ TEST FALLIDO (obtenidos: {len(cufes_xlsx)})")
        
        # Verificar que son los mismos CUFEs
        if set(cufes_txt) == set(cufes_xlsx):
            print("\n✅ BONUS: Ambos formatos retornan los mismos CUFEs")
        else:
            print("\n⚠️  Los CUFEs de .txt y .xlsx no coinciden")
    
    print("\n" + "="*70)
    print("✅ TESTS COMPLETADOS")
    print("="*70 + "\n")

if __name__ == "__main__":
    main()
